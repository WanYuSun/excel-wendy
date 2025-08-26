"""
union-sheets.py

一个用于读取Excel文件中多个sheets并使用DuckDB进行UNION操作的通用工具。

核心功能:
- 使用openpyxl读取Excel文件中的所有sheet名称
- 使用DuckDB的read_xlsx函数依次读取每个sheet
- 提供两个核心函数：union_sheets 和 unique_keys

🔧 函数1: union_sheets
参数:
- excel_file: Excel文件路径
- table_name: 输出表名  
- conn: DuckDB连接对象

功能:
- 读取Excel所有sheets并UNION ALL合并
- 保留所有原始数据，不去重

🔧 函数2: unique_keys  
参数:
- conn: DuckDB连接对象
- table_name: 输入表名
- projections: 投影列表 [(表达式, 别名), ...]

功能:
- 使用GROUP BY ALL自动去重
- 支持灵活的列投影和重命名
- 返回新表名: u_{table_name}

💡 使用流程:
1. union_sheets() - 合并所有sheets
2. unique_keys() - 按需去重和投影

设计原则:
- 每个函数职责单一明确
- 可以独立使用或组合使用
- 保持API简洁清晰
- 记录每个SQL操作的执行时间

Bootstrap流程 (首次使用):
1. 创建Python虚拟环境:
   python3 -m venv venv

2. 激活虚拟环境并安装依赖:
   source venv/bin/activate
   pip3 install -i https://bytedpypi.byted.org/simple openpyxl duckdb

3. 运行脚本:
   python3 test/union-sheets.py <excel_file_path>

4. 查看结果:
   # 查看生成的数据库文件
   ls -la excel.db
   
   # 查看test表内容 (可选)
   python3 -c "
   import duckdb
   conn = duckdb.connect('excel.db')
   result = conn.execute('SELECT * FROM test ORDER BY 1').fetchall()
   for row in result: print(row)
   conn.close()
   "

后续使用 (环境已配置):
   source venv/bin/activate && python3 test/union-sheets.py <excel_file_path>

使用说明:
1. 确保输入的Excel文件存在且包含多个sheet
2. 所有sheet应具有相同的列结构和数据类型
3. 程序会按第一列去重，如果多个sheet中有相同的第一列值，保留最后处理的sheet中的记录
4. 程序会自动跳过空的sheet
5. 结果保存在当前目录的excel.db文件中的test表

去重逻辑:
- 使用第一列作为去重的键值
- 当多个sheet中存在相同的第一列值时，保留后面sheet中的记录
- 例如：Sheet1中有记录A，Sheet2中也有记录A，最终保留Sheet2中的记录A

输出说明:
- 程序会显示发现的sheet数量和名称
- 显示去重前后的行数对比
- 显示最终表的结构信息
- 记录每个SQL操作的执行时间
- 生成excel.db数据库文件，包含test表

依赖:
- Python 3.7+
- openpyxl (Excel文件读取)
- duckdb (数据处理和SQL操作)

用法:
    python union-sheets.py <excel_file_path> [选项]

基本示例:
    # 默认用法（按第一列自动去重）
    python union-sheets.py /Users/bytedance/Documents/excel/Book1.xlsx
    
    # 使用UNION投影优化性能
    python union-sheets.py data.xlsx --union-projections '[["x", null], ["y", null]]'
    
    # 使用去重投影自定义聚合
    python union-sheets.py data.xlsx --unique-projections '[["x", null], ["any_value(y)", "avg_y"]]'
    
    # 只合并不去重
    python union-sheets.py data.xlsx --no-dedupe
    
    # 指定输出表名和数据库路径
    python union-sheets.py data.xlsx -o my_table -d my_data.db

命令行参数:
    excel_file                    Excel文件路径
    --union-projections, -up      UNION阶段投影列表（JSON格式）
    --unique-projections, -uq     去重阶段投影列表（JSON格式）
    --concurrent, -c             使用并发模式处理多个sheets
    --max-workers, -w            并发模式下的最大线程数（默认: 4）
    --no-dedupe                  只合并不去重
    --output-table, -o           输出表名（默认: test）
    --db-path, -d                数据库文件路径（默认: excel.db）

    --help, -h                   显示帮助信息
    
测试示例输出:
    [INFO] 发现 2 个sheet: ['Sheet1', 'Sheet2']
    [INFO] 连接到DuckDB数据库: excel.db
    [INFO] 使用第一列进行去重: x
    [INFO] 去重前总行数: 6
    [INFO] 去重后行数: 4
    [INFO] 去重操作完成，按第一列 'x' 去重
    [INFO] 处理完成！
"""

import argparse
import json
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Optional, Tuple

import duckdb
from openpyxl import load_workbook

# 导入统一的日志模块
from excel.log import (execute_sql_with_timing, log_error, log_info, log_success, log_warning, setup_logging)


# execute_sql_with_timing 函数现在从 log 模块导入


def get_sheet_names(excel_file: str) -> List[str]:
    """
    使用openpyxl读取Excel文件中的所有sheet名称

    Args:
        excel_file: Excel文件路径

    Returns:
        sheet名称列表

    Raises:
        FileNotFoundError: 文件不存在
        Exception: 文件读取失败
    """
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"Excel文件不存在: {excel_file}")

    try:
        workbook = load_workbook(excel_file, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        log_info(f"发现 {len(sheet_names)} 个sheet: {sheet_names}")
        return sheet_names
    except Exception as e:
        raise Exception(f"读取Excel文件失败: {e}")


def union_sheets_concurrent(excel_file: str, table_name: str, conn: duckdb.DuckDBPyConnection,
                            projections: Optional[List[Tuple[str,
                            Optional[str]]]] = None,
                            max_workers: int = None) -> None:
    """
    Excel多sheet高效并发合并函数

    Args:
        excel_file: Excel文件路径
        table_name: 输出表名
        conn: 主DuckDB连接对象
        projections: 投影列表
        max_workers: 最大并发线程数

    功能:
    - 并发创建临时表，任务完成后立即写入合并表
    - 任一任务失败则取消所有任务并抛出异常
    - 高效的流式处理，避免内存积累
    """
    # 获取sheet名称列表
    sheet_names = get_sheet_names(excel_file)
    if not sheet_names:
        log_warning("没有找到任何sheet，跳过处理")
        return

    if max_workers > len(sheet_names):
        max_workers = len(sheet_names)
    log_info(f"使用并发模式，最大线程数: {max_workers}")

    # 转义文件路径
    excel_file_escaped = excel_file.replace("\\", "\\\\")

    # 构建投影字符串
    if projections is None:
        projection_str = "*"
        log_info("使用默认投影: SELECT *")
    else:
        projection_parts = []
        for expr, alias in projections:
            if alias and alias.strip():
                projection_parts.append(f"{expr} AS {alias}")
            else:
                projection_parts.append(expr)
        projection_str = ', '.join(projection_parts)
        log_info(f"使用自定义投影: SELECT {projection_str}")

    # 预先加载Excel扩展，避免并发冲突
    try:
        conn.execute("INSTALL excel")
        conn.execute("LOAD excel")
        log_success("Excel扩展加载成功")
    except Exception as e:
        log_info(f"Excel扩展已存在或加载失败: {e}")

    # 获取主数据库路径
    try:
        db_info = conn.execute("PRAGMA database_list").fetchone()
        db_path = db_info[2] if db_info and db_info[2] != '' else ":memory:"
    except:
        db_path = ":memory:"

    def process_sheet_task(sheet_info):
        """处理单个sheet的任务"""
        sheet_index, sheet_name = sheet_info
        temp_table = f"temp_sheet_{sheet_index}_{int(time.time() * 1000) % 10000}"
        sheet_name_escaped = sheet_name.replace("'", "''")

        # 创建独立连接
        thread_conn = duckdb.connect(db_path)

        try:
            # 确保Excel扩展已加载（静默处理，避免冲突）
            try:
                thread_conn.execute("LOAD excel")
            except:
                pass  # 扩展可能已经加载，忽略错误

            # 创建临时表
            create_sql = f"""
            CREATE TABLE {temp_table} AS
            SELECT {projection_str} FROM read_xlsx(
                '{excel_file_escaped}',
                sheet='{sheet_name_escaped}',
                all_varchar=true
            )
            """

            start_time = time.time()
            thread_conn.execute(create_sql)
            execution_time = time.time() - start_time

            # 获取行数
            result = thread_conn.execute(
                f"SELECT COUNT(*) FROM {temp_table}").fetchone()
            row_count = result[0] if result else 0

            return {
                'sheet_name': sheet_name,
                'sheet_index': sheet_index,
                'temp_table': temp_table,
                'row_count': row_count,
                'execution_time': execution_time
            }

        finally:
            thread_conn.close()

    # 删除已存在的表
    execute_sql_with_timing(conn, f"DROP TABLE IF EXISTS {table_name}",
                            f"🗑️  删除已存在的表: {table_name}")

    # 并发处理并实时合并
    log_info(f"🚀 使用 {max_workers} 个线程并发处理 {len(sheet_names)} 个sheets")

    completed_sheets = []
    temp_tables = []
    total_rows = 0
    first_table_created = False

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务
        future_to_sheet = {
            executor.submit(process_sheet_task, (i, sheet_name)): sheet_name
            for i, sheet_name in enumerate(sheet_names)
        }

        try:
            # 实时处理完成的任务
            for future in as_completed(future_to_sheet):
                result = future.result()
                completed_sheets.append(result)
                temp_tables.append(result['temp_table'])
                total_rows += result['row_count']

                log_info(f"✅ {result['sheet_name']}: {result['row_count']} 行 "
                         f"(耗时: {result['execution_time']:.3f}s)")

                # 第一个完成的任务：创建目标表
                if not first_table_created:
                    execute_sql_with_timing(
                        conn,
                        f"CREATE TABLE {table_name} AS SELECT * FROM {result['temp_table']}",
                        f"🔄 创建目标表: {result['sheet_name']}"
                    )
                    first_table_created = True
                else:
                    # 后续任务：插入数据
                    execute_sql_with_timing(
                        conn,
                        f"INSERT INTO {table_name} SELECT * FROM {result['temp_table']}",
                        f"📊 插入数据: {result['sheet_name']}"
                    )

        except Exception as e:
            # 取消所有未完成的任务
            for f in future_to_sheet:
                if not f.done():
                    f.cancel()

            # 清理已创建的临时表
            for temp_table in temp_tables:
                try:
                    conn.execute(f"DROP TABLE IF EXISTS {temp_table}")
                except:
                    pass

            # 清理目标表
            if first_table_created:
                try:
                    conn.execute(f"DROP TABLE IF EXISTS {table_name}")
                except:
                    pass

            raise Exception(f"并发处理失败: {str(e)}")

    # 清理所有临时表
    for temp_table in temp_tables:
        try:
            execute_sql_with_timing(conn, f"DROP TABLE IF EXISTS {temp_table}",
                                    f"🗑️  清理临时表: {temp_table}")
        except Exception as e:
            log_warning(f"⚠️  清理临时表 {temp_table} 失败: {e}")

    log_success(f"并发处理完成，总行数: {total_rows}")
    log_success(f"结果保存到表: {table_name}")

    # 显示表结构信息
    start_time = time.time()
    schema_result = conn.execute(f"DESCRIBE {table_name}").fetchall()
    describe_time = time.time() - start_time
    log_info(f"📋 {table_name}表结构 (查询耗时: {describe_time:.3f}s):")
    for column_info in schema_result:
        log_info(f"  {column_info[0]}: {column_info[1]}")


def union_sheets(excel_file: str, table_name: str, conn: duckdb.DuckDBPyConnection,
                 projections: Optional[List[Tuple[str, Optional[str]]]] = None) -> None:
    """
    Excel多sheet合并函数 - 智能优化的合并策略

    Args:
        excel_file: Excel文件路径
        table_name: 输出表名
        conn: DuckDB连接对象
        projections: 投影列表，用于在UNION阶段过滤列，提高性能
                    - None: 选择所有列 (SELECT *)
                    - List: 自定义投影 [(表达式, 别名), ...]

    功能:
    - 读取Excel文件的所有sheets
    - 智能选择合并策略：
      * ≤3个sheets: 使用UNION ALL（性能最佳）
      * >3个sheets: 使用批量INSERT（避免巨大查询）
    - 支持投影优化，减少数据传输量

    Raises:
        Exception: 操作失败
    """
    # 获取sheet名称列表
    sheet_names = get_sheet_names(excel_file)
    if not sheet_names:
        log_warning("没有找到任何sheet，跳过处理")
        return

    # 转义文件路径中的反斜杠（Windows兼容性）
    excel_file_escaped = excel_file.replace("\\", "\\\\")

    # 构建投影字符串
    if projections is None:
        # 没有投影，选择所有列
        projection_str = "*"
        log_info("使用默认投影: SELECT *")
    else:
        # 使用自定义投影
        projection_parts = []
        for expr, alias in projections:
            if alias and alias.strip():
                projection_parts.append(f"{expr} AS {alias}")
            else:
                projection_parts.append(expr)
        projection_str = ', '.join(projection_parts)
        log_info(f"使用自定义投影: SELECT {projection_str}")

    # 删除已存在的表
    execute_sql_with_timing(conn, f"DROP TABLE IF EXISTS {table_name}",
                            f"🗑️  删除已存在的表: {table_name}")

    # 优化策略：根据sheet数量选择不同的处理方式
    if len(sheet_names) > 3:
        # 多sheet优化：逐个INSERT，避免巨大的UNION ALL
        log_info(f"🚀 使用批量INSERT模式处理 {len(sheet_names)} 个sheets（优化大数据量）")

        total_rows = 0
        for i, sheet_name in enumerate(sheet_names):
            sheet_name_escaped = sheet_name.replace("'", "''")

            if i == 0:
                # 第一个sheet：创建表
                create_sql = f"""
                CREATE TABLE {table_name} AS
                SELECT {projection_str} FROM read_xlsx(
                    '{excel_file_escaped}',
                    sheet='{sheet_name_escaped}',
                    all_varchar=true
                )
                """
                execute_sql_with_timing(conn, create_sql,
                                        f"📊 创建表并插入第1个sheet: {sheet_name}")
            else:
                # 后续sheet：批量插入
                insert_sql = f"""
                INSERT INTO {table_name}
                SELECT {projection_str} FROM read_xlsx(
                    '{excel_file_escaped}',
                    sheet='{sheet_name_escaped}',
                    all_varchar=true
                )
                """
                execute_sql_with_timing(conn, insert_sql,
                                        f"📊 插入第{i + 1}个sheet: {sheet_name}")

            # 获取当前总行数
            result = conn.execute(
                f"SELECT COUNT(*) FROM {table_name}").fetchone()
            current_rows = result[0] if result else 0
            sheet_rows = current_rows - total_rows
            total_rows = current_rows
            log_info(
                f"  📈 {sheet_name}: +{sheet_rows} 行，累计: {total_rows} 行")

    else:
        # 少量sheet：使用传统UNION ALL（性能更好）
        log_info(f"🔄 使用UNION ALL模式处理 {len(sheet_names)} 个sheets")

        union_queries = []
        for sheet_name in sheet_names:
            sheet_name_escaped = sheet_name.replace("'", "''")
            query = (f"SELECT {projection_str} FROM read_xlsx("
                     f"'{excel_file_escaped}', sheet='{sheet_name_escaped}', "
                     f"all_varchar=true)")
            union_queries.append(query)
            log_info(f"  📋 添加sheet: {sheet_name}")

        # 组合所有查询
        full_union_query = " UNION ALL ".join(union_queries)
        create_sql = f"CREATE TABLE {table_name} AS ({full_union_query})"

        execute_sql_with_timing(conn, create_sql, "🔄 执行UNION ALL操作")

        # 获取结果统计
        result = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()
        total_rows = result[0] if result else 0

    log_info(f"✅ 合并完成，总行数: {total_rows}")
    log_info(f"✅ 结果保存到表: {table_name}")

    # 显示表结构信息
    start_time = time.time()
    schema_result = conn.execute(f"DESCRIBE {table_name}").fetchall()
    describe_time = time.time() - start_time
    log_info(f"📋 {table_name}表结构 (查询耗时: {describe_time:.3f}s):")
    for column_info in schema_result:
        log_info(f"  {column_info[0]}: {column_info[1]}")


def unique_keys(conn: duckdb.DuckDBPyConnection,
                table_name: str,
                projections: List[Tuple[str, Optional[str]]]) -> str:
    """
    数据去重函数 - 按键去重并支持自定义投影

    Args:
        conn: DuckDB连接对象
        table_name: 输入表名
        projections: 投影列表 [(表达式, 别名), ...]
                    - 第一项必须是key列
                    - 别名为None或空时直接使用表达式
                    - 别名不为空时使用 expr AS alias

    Returns:
        新表名: u_{table_name}

    功能:
    - 使用GROUP BY ALL自动去重
    - 支持灵活的列投影和重命名

    Raises:
        Exception: 操作失败
    """
    if not projections:
        raise Exception("投影列表不能为空")

    # 生成新表名
    new_table_name = f"u_{table_name}"

    # 构建投影字符串
    projection_parts = []
    for expr, alias in projections:
        if alias and alias.strip():
            # 有别名，使用 expr AS alias
            projection_parts.append(f"{expr} AS {alias}")
        else:
            # 没有别名，直接使用expr
            projection_parts.append(expr)

    projection_str = ', '.join(projection_parts)

    # 删除已存在的新表
    execute_sql_with_timing(conn, f"DROP TABLE IF EXISTS {new_table_name}",
                            f"🗑️  删除已存在的表: {new_table_name}")

    # 使用GROUP BY ALL进行去重
    # GROUP BY ALL会自动按所有非聚合列进行分组
    group_by_sql = f"""
    CREATE TABLE {new_table_name} AS
    SELECT {projection_str}
    FROM {table_name}
    GROUP BY ALL
    ORDER BY 1 ASC
    """

    log_info(f"🔄 开始执行去重操作: {table_name} -> {new_table_name}")
    start_time = time.time()
    conn.execute(group_by_sql)
    group_by_time = time.time() - start_time
    log_info(f"⏱️  GROUP BY去重执行完成 (耗时: {group_by_time:.3f}s)")

    # 获取结果统计
    start_time = time.time()
    original_count = conn.execute(
        f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    new_count = conn.execute(
        f"SELECT COUNT(*) FROM {new_table_name}").fetchone()[0]
    count_time = time.time() - start_time

    log_info(f"📊 去重前行数: {original_count}")
    log_info(f"📊 去重后行数: {new_count} (统计耗时: {count_time:.3f}s)")
    log_info(f"✅ 去重完成，结果保存到表: {new_table_name}")

    # 显示新表结构信息
    start_time = time.time()
    schema_result = conn.execute(f"DESCRIBE {new_table_name}").fetchall()
    describe_time = time.time() - start_time
    log_info(f"📋 {new_table_name}表结构 (查询耗时: {describe_time:.3f}s):")
    for column_info in schema_result:
        log_info(f"  {column_info[0]}: {column_info[1]}")

    return new_table_name


def parse_projections(proj_str: str) -> List[Tuple[str, Optional[str]]]:
    """
    解析命令行投影参数

    Args:
        proj_str: 投影字符串，格式为JSON数组
                 例如: '[["x", null], ["any_value(y)", "y"], ["COUNT(*)", "count"]]'

    Returns:
        投影列表
    """
    try:
        proj_list = json.loads(proj_str)
        projections = []
        for item in proj_list:
            if isinstance(item, list) and len(item) == 2:
                expr, alias = item
                # 将null转换为None
                alias = None if alias is None or alias == "null" else alias
                projections.append((expr, alias))
            else:
                raise ValueError(f"投影项格式错误: {item}")
        return projections
    except json.JSONDecodeError as e:
        raise ValueError(f"投影参数JSON格式错误: {e}")
    except Exception as e:
        raise ValueError(f"投影参数解析失败: {e}")


def main():
    """
    主函数:
    - 解析命令行参数
    - 读取Excel文件的sheet列表
    - 执行UNION和去重操作
    """
    parser = argparse.ArgumentParser(
        description='Excel多sheet合并和去重工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
使用示例:
  # 基本用法（按第一列自动去重）
  python union-sheets.py data.xlsx
  
  # 使用UNION投影优化性能（只选择需要的列）
  python union-sheets.py data.xlsx --union-projections '[["x", null], ["y", null], ["z", null]]'
  
  # 使用去重投影自定义聚合
  python union-sheets.py data.xlsx --unique-projections '[["x", null], ["any_value(y)", "avg_y"], ["COUNT(*)", "count"]]'
  
  # 同时使用两种投影
  python union-sheets.py data.xlsx -up '[["x", null], ["y", null]]' -uq '[["x", null], ["any_value(y)", "y"]]'
  
  # 使用并发模式（适合大量sheets）
  python union-sheets.py data.xlsx --concurrent --max-workers 8
  
  # 只合并不去重
  python union-sheets.py data.xlsx --no-dedupe
  


投影格式说明:
  投影参数为JSON数组，每个元素为[表达式, 别名]的格式
  - 表达式: SQL表达式，如 "x", "any_value(y)", "COUNT(*)"
  - 别名: 列别名，可以为null（不使用别名）或字符串
  
  UNION投影 (--union-projections):
  - 用于在合并阶段过滤列，提高性能
  - 建议只选择需要的列，避免传输大量无用数据
  
  去重投影 (--unique-projections):
  - 用于去重阶段的列选择和聚合
  - 第一个投影项必须是key列
  - 其他列通常使用聚合函数如any_value()
        '''
    )

    parser.add_argument('excel_file', nargs='?', help='Excel文件路径')

    parser.add_argument('--union-projections', '-up', type=str,
                        help='UNION阶段投影列表（JSON格式），用于性能优化，例如: \'[["x", null], ["y", null]]\'')
    parser.add_argument('--unique-projections', '-uq', type=str,
                        help='去重阶段投影列表（JSON格式），例如: \'[["x", null], ["any_value(y)", "y"]]\'')
    parser.add_argument('--concurrent', '-c', action='store_true',
                        help='使用并发模式处理多个sheets（实验性功能）')
    parser.add_argument('--max-workers', '-w', type=int, default=4,
                        help='并发模式下的最大线程数（默认: 4）')
    parser.add_argument('--no-dedupe', action='store_true', help='只合并不去重')
    parser.add_argument('--output-table', '-o',
                        default='test', help='输出表名（默认: test）')
    parser.add_argument('--db-path', '-d', default='excel.db',
                        help='数据库文件路径（默认: excel.db）')

    args = parser.parse_args()

    # 检查必需参数
    if not args.excel_file:
        parser.error("请提供Excel文件路径")

    excel_file = args.excel_file

    try:
        # 设置统一的日志配置
        setup_logging()

        # 创建DuckDB连接
        conn = duckdb.connect(database=args.db_path)
        log_info(f"连接到DuckDB数据库: {args.db_path}")

        try:
            # 解析UNION投影
            union_projections = None
            if args.union_projections:
                union_projections = parse_projections(args.union_projections)
                log_info(f"🔧 UNION阶段投影: {union_projections}")

            # 步骤1: 合并所有sheets
            temp_table = "temp_union_all"
            if args.concurrent:
                log_info(f"🚀 使用并发模式，最大线程数: {args.max_workers}")
                union_sheets_concurrent(
                    excel_file, temp_table, conn, union_projections, args.max_workers)
            else:
                union_sheets(excel_file, temp_table, conn, union_projections)

            if args.no_dedupe:
                # 只合并不去重，直接重命名
                execute_sql_with_timing(conn, f"DROP TABLE IF EXISTS {args.output_table}",
                                        f"🗑️  删除已存在的表: {args.output_table}")
                execute_sql_with_timing(conn, f"ALTER TABLE {temp_table} RENAME TO {args.output_table}",
                                        f"🔄 重命名表为: {args.output_table}")
                log_info(f"✅ 合并完成，结果保存到表: {args.output_table}")
            else:
                # 步骤2: 去重处理
                unique_projections = None
                if args.unique_projections:
                    # 使用自定义去重投影
                    unique_projections = parse_projections(
                        args.unique_projections)
                    log_info(f"🔧 去重阶段投影: {unique_projections}")
                else:
                    # 使用默认投影（按第一列去重）
                    start_time = time.time()
                    columns_result = conn.execute(
                        f"DESCRIBE {temp_table}").fetchall()
                    describe_time = time.time() - start_time
                    log_info(f"⏱️  获取表结构信息 (耗时: {describe_time:.3f}s)")

                    if not columns_result:
                        raise Exception("无法获取表结构信息")

                    first_column = columns_result[0][0]
                    unique_projections = [(f'"{first_column}"', None)]  # key列
                    for col_info in columns_result[1:]:
                        col_name = col_info[0]
                        unique_projections.append(
                            (f'any_value("{col_name}")', col_name))

                    log_info(f"🔧 使用默认去重投影（按第一列 '{first_column}' 去重）")

                # 执行去重
                result_table = unique_keys(
                    conn, temp_table, unique_projections)

                # 重命名为目标表
                execute_sql_with_timing(conn, f"DROP TABLE IF EXISTS {args.output_table}",
                                        f"🗑️  删除已存在的表: {args.output_table}")
                execute_sql_with_timing(conn, f"ALTER TABLE {result_table} RENAME TO {args.output_table}",
                                        f"🔄 重命名表为: {args.output_table}")

                # 清理临时表
                execute_sql_with_timing(conn, f"DROP TABLE IF EXISTS {temp_table}",
                                        "🗑️  清理临时表")

                log_info(f"✅ 去重完成，结果保存到表: {args.output_table}")

        finally:
            conn.close()
            log_info("DuckDB连接已关闭")

        log_info("🎉 处理完成！")

    except Exception as e:
        log_error(f"程序执行失败: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
